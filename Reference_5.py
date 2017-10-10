import requests
import openpyxl

url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"

r = requests.get(url)
type(r.content)
f = open("hospital_ranking_focus_states.xlsx","wb")
f.write(r.content)
f.close()

wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")

for sheet_name in wb.get_sheet_names():
    print(sheet_name)
    
sheet_1 = wb.get_sheet_by_name("Hospital National Ranking")
sheet_2 = wb.get_sheet_by_name("Focus States")

i = 1
while sheet_1.cell(row = i, column = 1).value != None:
    print(sheet_1.cell(row = i,column = 1).value, "|", sheet_1.cell(row = i, column = 2).value)
    i += 1

i = k
while sheet_2.cell(row = k, column = 1).value != None:
    print(sheet_2.cell(row = k,column = 1).value, "|", sheet_2.cell(row = k, column = 2).value)
    i += k